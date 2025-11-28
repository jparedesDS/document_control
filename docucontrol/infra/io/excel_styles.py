from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.formatting.rule import Rule
from openpyxl.styles import Border, Font, PatternFill, Side
from openpyxl.styles.differential import DifferentialStyle


class MonitoringReportStyler:
    DATE_COLUMNS = ["Fecha", "Fecha Pedido", "Fecha Prevista", "Fecha Dev. Doc.", "Fecha Env. Doc.", "Fecha Doc."]
    TAB_COLORS = {
        "ALL DOC.": "6678AF",
        "ENVIADOS": "00D25F",
        "DEVOLUCIONES": "FFA19A",
        "CRÍTICOS": "DBB054",
        "CRÍTICOS +15d": "FF7F50",
        "SIN ENVIAR": "FFFF66",
        "STATUS GLOBAL": "B1E1B9",
    }
    ESTADO_COLORES = {
        "Rechazado": "FFA19A",
        "Com. Menores": "FFE5AD",
        "Com. Mayores": "DBB054",
        "Comentado": "F79646",
        "Enviado": "B1E1B9",
        "Sin Enviar": "FFFFAB",
        "Información": "FFFF46",
        "HOLD": "FF0909",
        "Aprobado": "00D25F",
    }
    RESPONSABLES_COLORES = {
        "SS": "262626",
        "CCH": "1F1F1F",
        "JM": "2C8A6A",
        "JV": "006B95",
        "EC": "B31274",
        "ES": "5A0DA0",
        "JP": "00458F",
        "AC": "3F0075",
        "LB": "176DD1",
        "RM": "228B22",
        "RP": "1B365D",
        "EC/SS": "1F7A1F",
    }

    def apply(self, workbook_path: Path) -> None:
        wb = load_workbook(workbook_path)
        for sheet in wb.worksheets:
            self._style_sheet(sheet)
            self._format_dates(sheet)
            self._auto_filter(sheet)
            self._auto_width(sheet)
            self._color_estado(sheet)
            self._color_responsables(sheet)
            self._highlight_columns(sheet)
            if sheet.title in self.TAB_COLORS:
                sheet.sheet_properties.tabColor = self.TAB_COLORS[sheet.title]
            if sheet.title == "DEVOLUCIONES":
                self._resaltar_dias_devolucion(sheet)
        self._add_status_chart(wb)
        wb.save(workbook_path)

    def _style_sheet(self, sheet) -> None:
        fill_light = PatternFill(start_color="D4DCF4", end_color="D4DCF4", fill_type="solid")
        fill_dark = PatternFill(start_color="6678AF", end_color="6678AF", fill_type="solid")
        font_white = Font(color="FFFFFF", bold=True)
        border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))

        for cell in sheet[1]:
            cell.fill = fill_dark
            cell.font = font_white
            cell.border = border

        for row in sheet.iter_rows(min_row=2):
            for cell in row:
                cell.fill = fill_light
                cell.border = border

        sheet.freeze_panes = sheet["B2"]

    def _format_dates(self, sheet) -> None:
        header_map = {cell.value: idx + 1 for idx, cell in enumerate(sheet[1])}
        for col_name in self.DATE_COLUMNS:
            if col_name in header_map:
                col_idx = header_map[col_name]
                for column in sheet.iter_cols(min_col=col_idx, max_col=col_idx, min_row=2):
                    for cell in column:
                        if cell.value:
                            cell.number_format = "DD/MM/YYYY"

    @staticmethod
    def _auto_filter(sheet) -> None:
        sheet.auto_filter.ref = sheet.dimensions

    @staticmethod
    def _auto_width(sheet) -> None:
        for col_cells in sheet.iter_cols(min_row=1, max_row=sheet.max_row):
            header = col_cells[0]
            max_length = len(str(header.value)) if header.value else 0
            for cell in col_cells[1:]:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            sheet.column_dimensions[header.column_letter].width = max_length + 5

    def _color_estado(self, sheet) -> None:
        col_idx = self._find_column(sheet, "Estado")
        if not col_idx:
            return
        for row in sheet.iter_rows(min_row=2):
            estado = row[col_idx - 1].value
            if estado in self.ESTADO_COLORES:
                row[col_idx - 1].fill = PatternFill(
                    start_color=self.ESTADO_COLORES[estado],
                    end_color=self.ESTADO_COLORES[estado],
                    fill_type="solid",
                )
                if estado in {"HOLD", "Aprobado"}:
                    row[col_idx - 1].font = Font(color="FF0000" if estado == "HOLD" else "000000", bold=True)

    def _color_responsables(self, sheet) -> None:
        indices = [
            idx for idx in (self._find_column(sheet, "Responsable"), self._find_column(sheet, "Repsonsable")) if idx
        ]
        for idx in indices:
            for row in sheet.iter_rows(min_row=2):
                val = row[idx - 1].value
                if val in self.RESPONSABLES_COLORES:
                    row[idx - 1].font = Font(color=self.RESPONSABLES_COLORES[val], bold=True)

    def _highlight_columns(self, sheet) -> None:
        crit_idx = self._find_column(sheet, "Crítico")
        info_idx = self._find_column(sheet, "Info/Review")
        dias_env_idx = self._find_column(sheet, "Días Envío")
        dias_dev_idx = self._find_column(sheet, "Días Devolución")

        if crit_idx:
            for row in sheet.iter_rows(min_row=2):
                if row[crit_idx - 1].value == "Sí":
                    row[crit_idx - 1].font = Font(color="FF0000", bold=True)

        if info_idx:
            for row in sheet.iter_rows(min_row=2):
                val = row[info_idx - 1].value
                if val == "R":
                    row[info_idx - 1].font = Font(color="FF0000", bold=True)
                elif val == "I":
                    row[info_idx - 1].font = Font(color="4D4D4D", bold=True)

        if dias_env_idx:
            for row in sheet.iter_rows(min_row=2):
                if row[dias_env_idx - 1].value == 15:
                    row[dias_env_idx - 1].font = Font(color="4D4D4D", bold=True)

        if dias_dev_idx:
            for row in sheet.iter_rows(min_row=2):
                val = row[dias_dev_idx - 1].value
                if isinstance(val, (int, float)) and val > 15:
                    row[dias_dev_idx - 1].font = Font(color="FF0000", bold=True)

    def _resaltar_dias_devolucion(self, sheet) -> None:
        col_idx = self._find_column(sheet, "Días Devolución")
        if not col_idx:
            return
        diff = DifferentialStyle(fill=PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"))
        col_letter = sheet.cell(row=2, column=col_idx).column_letter
        formula = f"=${col_letter}2>15"
        rango = f"A2:{sheet.cell(row=sheet.max_row, column=sheet.max_column).coordinate}"
        sheet.conditional_formatting.add(rango, Rule(type="expression", formula=[formula], dxf=diff))

    def _add_status_chart(self, wb) -> None:
        if "STATUS GLOBAL" not in wb.sheetnames:
            return
        ws = wb["STATUS GLOBAL"]
        headers = [cell.value for cell in ws[1]]
        estado_cols = ["Aprobado", "Com. Mayores", "Com. Menores", "Enviado", "Rechazado", "Sin Enviar"]
        indices = [headers.index(col) + 1 for col in estado_cols if col in headers]
        if not indices:
            return
        min_col, max_col = min(indices), max(indices)
        min_row, max_row = 2, ws.max_row
        data = Reference(ws, min_col=min_col, max_col=max_col, min_row=1, max_row=max_row)
        pedidos = Reference(ws, min_col=1, max_col=1, min_row=min_row, max_row=max_row)

        chart = BarChart()
        chart.type = "col"
        chart.title = "Estado por Pedido"
        chart.style = 12
        chart.grouping = "stacked"
        chart.overlap = 100
        chart.y_axis.title = "Nº Documentos"
        chart.x_axis.title = "Nº Pedido"
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(pedidos)
        colors = ["00B350", "C59B3F", "FFCF7F", "5566A0", "FF8273", "FFEF7F"]
        for serie, color in zip(chart.series, colors):
            serie.graphicalProperties = GraphicalProperties(solidFill=color)
        chart.height = 16
        chart.width = 29
        ws.add_chart(chart, "J3")

    @staticmethod
    def _find_column(sheet, header: str) -> int | None:
        for idx, cell in enumerate(sheet[1], start=1):
            if cell.value == header:
                return idx
        return None

