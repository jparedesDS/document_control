import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Font
from openpyxl.utils.dataframe import dataframe_to_rows


def aplicar_estilos_y_guardar_excel(df, filename):
    wb = Workbook()
    ws = wb.active

    cell_filling_blue_light = PatternFill(start_color="D4DCF4", end_color="D4DCF4", fill_type="solid")
    cell_filling = PatternFill(start_color="6678AF", end_color="6678AF", fill_type="solid")
    border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))
    font_white = Font(color="FFFFFF", bold=True)
    font_black = Font(color="000000")

    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True)):
        ws.append(row)
        for c_idx, _ in enumerate(row, 1):
            cell = ws.cell(row=r_idx + 1, column=c_idx)
            cell.border = border
            if r_idx == 0:
                cell.fill = cell_filling
                cell.font = font_white
            else:
                cell.fill = cell_filling_blue_light
                cell.font = font_black

    for column in ws.columns:
        max_length = max(len(str(cell.value)) if cell.value else 0 for cell in column)
        ws.column_dimensions[column[0].column_letter].width = max_length + 2

    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = "B2"
    wb.save(filename)


def aplicar_estilos_html(df):
    styles = {
        "fecha": "background-color: #D4DCF4; text-align: left; font-size: 14px;",
        "header": "background-color: #6678AF; color: #FFFFFF; text-align: left; font-size: 14px;",
        "cell_even": "background-color: #D4DCF4; text-align: left; font-size: 14px;",
    }

    def style_specific_cell(val):
        if isinstance(val, pd.Timestamp):
            return styles["fecha"]
        return styles["cell_even"]

    def apply_conditional_styles(val):
        mapping = {
            "Rechazado": "background-color: #FFA19A;",
            "Com. Menores": "background-color: #FFE5AD;",
            "Com. Mayores": "background-color: #DBB054;",
            "Comentado": "background-color: #F79646;",
            "Aprobado": "background-color: #00D25F;",
            "Eliminado": "background-color: #FF0000;",
        }
        if val in mapping:
            return f"color: #000000; font-weight: bold; {mapping[val]} font-size: 14px;"
        return "text-align: left; font-size: 14px;"

    header_style = [
        {
            "selector": "th",
            "props": [
                ("background-color", "#6678AF"),
                ("color", "#FFFFFF"),
                ("text-align", "center"),
                ("font-size", "14px"),
                ("font-weight", "bold"),
            ],
        }
    ]

    styled = df.style.map(style_specific_cell).map(apply_conditional_styles).set_table_styles(header_style)
    return styled.to_html(index=False)


def aplicar_estilo_info(df):
    estilo_celdas = "background-color: #D4DCF4; text-align: left; font-size: 14px;"
    estilo_header = [
        {
            "selector": "th",
            "props": [
                ("background-color", "#6678AF"),
                ("color", "#FFFFFF"),
                ("text-align", "center"),
                ("font-size", "15px"),
                ("font-weight", "bold"),
            ],
        }
    ]

    styled = df.style.map(lambda _: estilo_celdas).set_table_styles(estilo_header)
    return styled.to_html(index=False)

